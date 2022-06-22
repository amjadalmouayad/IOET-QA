import pandas as pd
import openpyxl
import datetime

import pytest

try:
    listTest = pd.read_excel("input.xlsx")
except FileNotFoundError as exception:
    pytest.exit("\n File not found and test failed | Please insert the input data for the test case in input.xlsx\n "
                "For more information please open Readme file")

num_test = len(listTest)
columns = len(listTest.columns)
arrListTest = listTest.to_numpy()


def update_input_file():
    filename = "./input.xlsx"
    wb = openpyxl.load_workbook(filename)
    sheet = wb['test data']
    sheet.delete_rows(2, 1)
    wb.save("./input.xlsx")


def update_output_file():
    now = datetime.datetime.now()
    list_output_test = pd.read_excel("./output.xlsx")
    num_output_test = len(list_output_test) + 2
    list_output = openpyxl.load_workbook('./output.xlsx')
    sheet = list_output.active
    for c in range(1, 5):
        # print("value " + str(c) + ": " + arrListTest[0][c-1])
        sheet.cell(row=num_output_test, column=c).value = arrListTest[0][c - 1]
    sheet.cell(row=num_output_test, column=5).value = now.strftime('%Y-%m-%d %H:%M:%S')
    list_output.save('./output.xlsx')
